from openpyxl import load_workbook


#NomeArquivo = input('Digite o nome do arquivo: ')

Arquivo = load_workbook ('caminho\\.xlsx')

Total = len(Arquivo.sheetnames)

ListaDeSheet = []

for a in range(0, Total, 1):
    if Arquivo.worksheets[a].sheet_state == 'hidden':
        pass
    else:
        ListaDeSheet.append(Arquivo.worksheets[a].title)

for x in ListaDeSheet:
    print(x)

# input('Digite enter pra continuar.')
